from fastapi import FastAPI, File, UploadFile, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
import logging
import os
from pathlib import Path

from app.models.schemas import (
    UploadResponse, AnalysisRequest, AnalysisResponse, 
    ExportRequest, ErrorResponse, ColumnMapping
)
from app.models.config import AppConfig
from app.services.simple_excel_service import SimpleExcelService
from app.services.simple_analysis_service import SimpleAnalysisService
from app.services.simple_export_service import SimpleExportService
from app.utils.file_utils import read_excel_file, get_sample_data, is_valid_excel_file
from app.utils.config_utils import ConfigManager, ResultManager

# 設定の読み込み（ログ設定より前に実行）
config = AppConfig.load_from_file()
config.ensure_directories()

# ログ設定
# 環境変数に基づいてログ設定を調整
log_level = os.getenv("LOG_LEVEL", "INFO")
environment = os.getenv("ENVIRONMENT", "development")

# ログハンドラーの設定
handlers = [logging.StreamHandler()]

# 本番環境以外ではファイルログも有効にする
if environment != "production":
    try:
        os.makedirs('logs', exist_ok=True)
        handlers.append(logging.FileHandler('logs/app.log', encoding='utf-8'))
    except Exception as e:
        # ファイルログが作成できない場合はストリームログのみ
        print(f"Warning: Could not create log file: {e}")

logging.basicConfig(
    level=getattr(logging, log_level.upper(), logging.INFO),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=handlers
)
logger = logging.getLogger(__name__)

# FastAPIアプリケーションの作成
app = FastAPI(
    title="Clustering Map API",
    description="Excelアンケート結果からクラスタリングマップを生成するAPI",
    version="0.1.0"
)

# CORS設定
cors_origins = os.getenv("CORS_ORIGINS", "http://localhost:3000,http://127.0.0.1:3000").split(",")
# 本番環境ではフロントエンドのURLを追加
if os.getenv("ENVIRONMENT") == "production" or os.getenv("VERCEL"):
    # Vercel環境の場合、同じドメインからのアクセスを許可
    vercel_url = os.getenv("VERCEL_URL", "")
    if vercel_url:
        cors_origins.extend([
            f"https://{vercel_url}",
            f"https://{vercel_url}/"
        ])
    # Render環境のURLも保持（後方互換性のため）
    cors_origins.extend([
        "https://clustering-map-frontend.onrender.com",
        "https://clustering-map-frontend.onrender.com/"
    ])

logger.info(f"CORS origins: {cors_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# サービスの初期化
excel_service = SimpleExcelService()
analysis_service = SimpleAnalysisService()
export_service = SimpleExportService()
config_manager = ConfigManager()
result_manager = ResultManager()

# 静的ファイルの配信
if os.path.exists("frontend/dist"):
    app.mount("/static", StaticFiles(directory="frontend/dist"), name="static")


@app.get("/")
async def root():
    """ルートエンドポイント"""
    return {"message": "Clustering Map API", "version": "0.1.0"}

@app.get("/health")
async def health_check():
    """ヘルスチェックエンドポイント"""
    return {
        "status": "healthy",
        "message": "Clustering Map API is running",
        "version": "0.1.0",
        "timestamp": datetime.now().isoformat()
    }




@app.get("/template")
async def download_template():
    """テンプレートファイルをダウンロード"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # 実際のビジネス文脈のサンプルデータを作成
        sample_data = [
            '22時以降の残業を禁止にして欲しいです。夜に連絡がくるのでワークライフバランスが保てません。',
            'チームの仲間がとても協力的で、困った時には助け合える環境です。上司も理解があります。',
            'スキルアップのための研修制度が充実していて、キャリア成長を実感できます。',
            '給与や待遇面で満足しており、ボーナスも期待できます。昇進の機会も多いです。',
            '会社の業績が好調で、売上が前年比で20%向上しました。目標を達成できて嬉しいです。',
            '職場環境が快適で、オフィスの設備も整っています。働きやすい環境です。',
            'プロジェクトの責任が重く、プレッシャーを感じることがあります。',
            '会社の文化や価値観に共感でき、働きがいを感じています。',
            '残業が多く、休暇が取りにくい状況が続いています。改善が必要です。',
            '夜中や休日に緊急の連絡が来ることがあり、プライベートの時間が取れません。'
        ]
        
        # ワークブックを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "アンケート結果"
        
        # ヘッダーを設定
        ws['A1'] = '自由記述'
        
        # ヘッダーのスタイル設定
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        
        # サンプルデータを追加
        for i, data in enumerate(sample_data, start=2):
            ws[f'A{i}'] = data
        
        # 列幅を調整
        ws.column_dimensions['A'].width = 80
        
        # データ行のスタイル設定
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # メモリ上に保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=clustering_map_template.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Template generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"テンプレートの生成に失敗しました: {str(e)}")


@app.post("/upload", response_model=UploadResponse)
async def upload_excel(file: UploadFile = File(...)):
    """Excelファイルをアップロードして前処理"""
    try:
        logger.info(f"Upload request received: filename={file.filename}, content_type={file.content_type}")
        # ファイル形式の検証
        if not file.filename or not is_valid_excel_file(file.filename):
            raise HTTPException(
                status_code=400,
                detail="サポートされていないファイル形式です。.xlsxまたは.xlsファイルをアップロードしてください。"
            )
        
        # ファイルサイズの検証
        content = await file.read()
        logger.info(f"File size: {len(content)} bytes")
        if len(content) > config.max_file_size:
            raise HTTPException(
                status_code=400,
                detail=f"ファイルサイズが大きすぎます。最大{config.max_file_size // (1024*1024)}MBまでです。"
            )
        
        # 一時ファイルに保存
        from app.utils.file_utils import create_temp_file, cleanup_temp_file
        temp_file_path = create_temp_file(content, '.xlsx')
        
        try:
            # Excelファイルを読み込み
            logger.info(f"Reading Excel file: {temp_file_path}")
            df = read_excel_file(temp_file_path)
            logger.info(f"Excel file loaded: {len(df)} rows, {len(df.columns)} columns")
            
            # 行数制限の検証
            if len(df) > config.max_rows:
                raise HTTPException(
                    status_code=400,
                    detail=f"データ行数が多すぎます。最大{config.max_rows}行までです。"
                )
            
            # サンプルデータを取得
            logger.info("Generating sample data...")
            sample_data = get_sample_data(df, 5)
            
            # タグ候補を生成
            logger.info("Generating tag candidates...")
            tag_candidates = excel_service.generate_tag_candidates(df)
            logger.info(f"Generated {len(tag_candidates)} tag candidates")
            
            logger.info("Upload processing completed successfully")
            return UploadResponse(
                success=True,
                message="ファイルのアップロードと前処理が完了しました。",
                columns=list(df.columns),
                sample_data=sample_data,
                tag_candidates=tag_candidates
            )
        
        finally:
            # 一時ファイルを削除
            cleanup_temp_file(temp_file_path)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"ファイルの処理中にエラーが発生しました: {str(e)}")


@app.post("/analyze", response_model=AnalysisResponse)
async def analyze_data(request: AnalysisRequest):
    """データの解析を実行"""
    try:
        # リクエスト内容をログに出力
        logger.info(f"Analysis request received: cluster_method={request.cluster_method}, shape_mask_path={request.shape_mask_path}")
        logger.info(f"Request dict: {request.model_dump()}")
        
        # 解析を実行
        result = analysis_service.analyze_data(request)
        
        return AnalysisResponse(
            success=True,
            message="解析が完了しました。",
            data_points=result["data_points"],
            clusters=result["clusters"],
            tags=result["tags"],
            config=result.get("config", {})
        )
    
    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"解析中にエラーが発生しました: {str(e)}")


@app.get("/export/pdf")
async def export_pdf():
    """PDFエクスポート（Vercelでは無効化）"""
    raise HTTPException(
        status_code=503, 
        detail="PDF export is not available on Vercel Serverless Functions due to package size constraints"
    )


@app.get("/export/png")
async def export_png():
    """PNGエクスポート（Vercelでは無効化）"""
    raise HTTPException(
        status_code=503, 
        detail="PNG export is not available on Vercel Serverless Functions due to package size constraints"
    )


@app.get("/tags")
async def get_tags():
    """タグ辞書を取得"""
    try:
        tags = excel_service.get_tag_rules()
        return {"success": True, "tags": tags}
    except Exception as e:
        logger.error(f"Get tags failed: {e}")
        raise HTTPException(status_code=500, detail=f"タグ辞書の取得中にエラーが発生しました: {str(e)}")


@app.post("/tags")
async def update_tags(tags: dict):
    """タグ辞書を更新"""
    try:
        excel_service.update_tag_rules(tags)
        return {"success": True, "message": "タグ辞書が更新されました。"}
    except Exception as e:
        logger.error(f"Update tags failed: {e}")
        raise HTTPException(status_code=500, detail=f"タグ辞書の更新中にエラーが発生しました: {str(e)}")


@app.get("/configs")
async def get_configs():
    """保存された設定一覧を取得"""
    try:
        configs = config_manager.list_configs()
        return {"success": True, "configs": configs}
    except Exception as e:
        logger.error(f"Get configs failed: {e}")
        raise HTTPException(status_code=500, detail=f"設定一覧の取得中にエラーが発生しました: {str(e)}")


@app.post("/configs")
async def save_config(config: dict):
    """設定を保存"""
    try:
        config_name = config.get("name")
        config_data = config.get("config")
        
        if not config_data:
            raise HTTPException(status_code=400, detail="設定データが提供されていません")
        
        saved_path = config_manager.save_analysis_config(config_data, config_name)
        return {"success": True, "message": "設定が保存されました。", "path": saved_path}
    except Exception as e:
        logger.error(f"Save config failed: {e}")
        raise HTTPException(status_code=500, detail=f"設定の保存中にエラーが発生しました: {str(e)}")


@app.get("/configs/{config_name}")
async def get_config(config_name: str):
    """設定を取得"""
    try:
        config_data = config_manager.load_analysis_config(config_name)
        if config_data is None:
            raise HTTPException(status_code=404, detail="設定が見つかりません")
        
        return {"success": True, "config": config_data}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get config failed: {e}")
        raise HTTPException(status_code=500, detail=f"設定の取得中にエラーが発生しました: {str(e)}")


@app.delete("/configs/{config_name}")
async def delete_config(config_name: str):
    """設定を削除"""
    try:
        success = config_manager.delete_config(config_name)
        if not success:
            raise HTTPException(status_code=404, detail="設定が見つかりません")
        
        return {"success": True, "message": "設定が削除されました。"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete config failed: {e}")
        raise HTTPException(status_code=500, detail=f"設定の削除中にエラーが発生しました: {str(e)}")


@app.get("/results")
async def get_results():
    """保存された結果一覧を取得"""
    try:
        results = result_manager.list_results()
        return {"success": True, "results": results}
    except Exception as e:
        logger.error(f"Get results failed: {e}")
        raise HTTPException(status_code=500, detail=f"結果一覧の取得中にエラーが発生しました: {str(e)}")


@app.get("/results/{result_name}")
async def get_result(result_name: str):
    """結果を取得"""
    try:
        result_data = result_manager.load_analysis_result(result_name)
        if result_data is None:
            raise HTTPException(status_code=404, detail="結果が見つかりません")
        
        return {"success": True, "result": result_data}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get result failed: {e}")
        raise HTTPException(status_code=500, detail=f"結果の取得中にエラーが発生しました: {str(e)}")


@app.delete("/results/{result_name}")
async def delete_result(result_name: str):
    """結果を削除"""
    try:
        success = result_manager.delete_result(result_name)
        if not success:
            raise HTTPException(status_code=404, detail="結果が見つかりません")
        
        return {"success": True, "message": "結果が削除されました。"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete result failed: {e}")
        raise HTTPException(status_code=500, detail=f"結果の削除中にエラーが発生しました: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
