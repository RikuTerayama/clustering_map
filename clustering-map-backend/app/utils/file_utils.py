import os
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


def read_excel_file(file_path: str):
    """Excelファイルを読み込む（openpyxlのみ使用）"""
    try:
        from openpyxl import load_workbook
        
        # openpyxlでExcelファイルを読み込み
        workbook = load_workbook(file_path, read_only=True)
        worksheet = workbook.active
        
        # データを読み込み
        data = []
        headers = []
        
        # ヘッダー行を取得
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value else f"Column_{len(headers)+1}")
        
        # データ行を取得
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # 空行をスキップ
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
        
        workbook.close()
        
        logger.info(f"Excel file loaded successfully: {len(data)} rows, {len(headers)} columns")
        
        # pandas DataFrame風のオブジェクトを作成
        class SimpleDataFrame:
            def __init__(self, data, columns):
                self.data = data
                self.columns = columns
                self._len = len(data)
            
            def __len__(self):
                return self._len
            
            def head(self, n=5):
                return SimpleDataFrame(self.data[:n], self.columns)
            
            def to_dict(self, orient='records'):
                if orient == 'records':
                    return self.data
                return {col: [row.get(col) for row in self.data] for col in self.columns}
        
        return SimpleDataFrame(data, headers)
        
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        raise ValueError(f"Excelファイルの読み込みに失敗しました: {e}")


def validate_excel_columns(df, required_columns: List[str]) -> bool:
    """Excelファイルの列を検証"""
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"必要な列が見つかりません: {missing_columns}")
    return True


def get_sample_data(df, n_rows: int = 5) -> List[Dict[str, Any]]:
    """サンプルデータを取得"""
    sample_df = df.head(n_rows)
    return sample_df.to_dict('records')


def save_results(data: Dict[str, Any], file_path: str) -> None:
    """結果をJSONファイルに保存"""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    import json
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)


def load_results(file_path: str) -> Dict[str, Any]:
    """結果をJSONファイルから読み込み"""
    import json
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_file_extension(file_path: str) -> str:
    """ファイル拡張子を取得"""
    return Path(file_path).suffix.lower()


def is_valid_excel_file(file_path: str) -> bool:
    """有効なExcelファイルかチェック"""
    valid_extensions = ['.xlsx', '.xls']
    return get_file_extension(file_path) in valid_extensions


def create_temp_file(content: bytes, suffix: str = '.xlsx') -> str:
    """一時ファイルを作成"""
    import tempfile
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_file.write(content)
    temp_file.close()
    return temp_file.name


def cleanup_temp_file(file_path: str) -> None:
    """一時ファイルを削除"""
    try:
        if os.path.exists(file_path):
            os.unlink(file_path)
    except Exception as e:
        logger.warning(f"Failed to cleanup temp file {file_path}: {e}")
